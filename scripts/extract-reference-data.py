#!/usr/bin/env python3
"""Extract the audited PRAS-TAMRA calculation inputs from the supplied workbooks.

The generated TypeScript file is committed so the browser calculation engine does
not need Excel at runtime.  Run this script only when the source workbooks change.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE_ROOT = Path("/workspace/scratch/702220e1046e/work/excel-source")
OUTPUT = Path(__file__).resolve().parents[1] / "lib/simulator/reference-data.generated.ts"

APPLIANCE_NAMES = {
    "MOBILE_IT": "노트북·태블릿·휴대폰충전",
    "GAME_CONSOLE": "게임콘솔",
    "DISHWASHER": "식기세척기",
    "FOOD_WASTE_PROCESSOR": "음식물처리기",
    "WASHER": "세탁기",
    "CLOTHES_DRYER": "의류건조기",
    "CLOTHING_CARE": "의류관리기",
    "ROBOT_VACUUM": "로봇청소기",
    "CORDLESS_VACUUM": "무선청소기·충전",
    "IRON": "다리미·스팀다리미",
    "LIVING_ROOM_AC": "에어컨_거실",
    "HEAT_PUMP_HEATING": "히트펌프난방",
    "BOILER_CIRCULATION_PUMP": "보일러순환펌프",
}

SEASON_MAP = {"봄가을": "SHOULDER", "여름": "SUMMER", "겨울": "WINTER"}
WEEKDAY_MAP = {"토": "SATURDAY", "일": "HOLIDAY"}
DAY_TYPE_OVERRIDES = {
    # KEPCO EV tariff excludes temporary public holidays from holiday pricing.
    "2025/01/27": "WEEKDAY",
    # Statutory holidays present in the 2026 YTD event rows.
    "2026/02/17": "HOLIDAY",
    "2026/02/18": "HOLIDAY",
    "2026/05/05": "HOLIDAY",
}


def compact(value: float) -> float:
    return round(float(value), 9)


def read_profiles():
    residential_path = SOURCE_ROOT / "탐라는 전기예보 시뮬레이션 계산_0%.xlsx"
    ws = load_workbook(residential_path, data_only=True)["계산시트_주말반영"]
    residential = {
        SEASON_MAP[ws.cell(row, 1).value]: [compact(ws.cell(row, col).value) for col in range(2, 26)]
        for row in (2, 3, 4)
    }

    result = {"RESIDENTIAL_TOU": residential}
    ev_specs = (
        ("EV_SLOW_LOW_VOLTAGE", "*완속_10가지*.xlsx"),
        ("EV_FAST_HIGH_VOLTAGE", "*급속1h충전.xlsx"),
    )
    for code, pattern in ev_specs:
        path = next(SOURCE_ROOT.glob(pattern))
        ws = load_workbook(path, data_only=True)["사이즈"]
        result[code] = {
            season: [compact(ws.cell(row, col).value) for col in range(2, 26)]
            for row, season in ((10, "SHOULDER"), (11, "SUMMER"), (12, "WINTER"))
        }
    return result


def read_appliance_profiles():
    profiles = {}
    for path in sorted(SOURCE_ROOT.glob("*가전 부하곡선 시나리오_(*).xlsx")):
        name = path.stem
        season_ko = next(key for key in SEASON_MAP if key in name)
        day_group = "WEEKEND" if "주말" in name else "WEEKDAY"
        workbook = load_workbook(path, data_only=True)
        sheet = workbook[workbook.sheetnames[2]]
        header_to_col = {sheet.cell(1, col).value: col for col in range(3, 44)}
        key = f"{SEASON_MAP[season_ko]}_{day_group}"
        profiles[key] = {
            code: [compact(sheet.cell(row, header_to_col[label]).value or 0) for row in range(2, 26)]
            for code, label in APPLIANCE_NAMES.items()
        }
    return profiles


def read_events():
    base_path = SOURCE_ROOT / "탐라는 전기예보 시뮬레이션 계산_0%.xlsx"
    shifted_path = SOURCE_ROOT / "탐라는 전기예보 시뮬레이션 계산_50%.xlsx"
    base = load_workbook(base_path, data_only=True)["계산시트_주말반영"]
    shifted = load_workbook(shifted_path, data_only=True)["계산시트_주말반영"]

    master_path = SOURCE_ROOT / "260623_판매수입 변화 계산용(주택용,전기차,일반용,산업용)_순부 수정.xlsx"
    master = load_workbook(master_path, data_only=True)["주택용TOU_수요이전0%"]
    master_day_type = {}
    for row in range(3, 59):
        master_day_type[master.cell(row, 1).value] = {
            "평일": "WEEKDAY",
            "토": "SATURDAY",
            "공휴일": "HOLIDAY",
        }[master.cell(row, 4).value]

    events = {"2024": [], "2025": [], "2026": []}
    for row in range(9, 128):
        raw_date = base.cell(row, 1).value
        if not isinstance(raw_date, str):
            continue
        try:
            date = datetime.strptime(raw_date, "%Y/%m/%d")
        except ValueError:
            continue
        if date.year not in (2024, 2025, 2026):
            continue

        smp = [compact(base.cell(row, col).value) for col in range(4, 28)]
        manual_hours = [
            hour
            for hour, col in enumerate(range(29, 53))
            if isinstance(shifted.cell(row, col).value, (int, float)) and shifted.cell(row, col).value > 1e-12
        ]
        # 2024/2025 workbooks contain the actual scenario windows.  The 2026
        # YTD rows do not, so use the published SMP<=0 candidate window.
        actual_hours = manual_hours or [
            hour for hour, value in enumerate(smp)
            if 9 <= hour <= 15 and value <= 0
        ]
        if not actual_hours:
            continue

        season_ko = base.cell(row, 2).value
        if season_ko not in SEASON_MAP:
            season_ko = "여름" if date.month in (6, 7, 8) else "겨울" if date.month in (11, 12, 1, 2) else "봄가을"
        weekday_ko = base.cell(row, 3).value
        calendar_day_type = "SATURDAY" if date.weekday() == 5 else "HOLIDAY" if date.weekday() == 6 else "WEEKDAY"
        day_type = DAY_TYPE_OVERRIDES.get(
            raw_date,
            master_day_type.get(raw_date, WEEKDAY_MAP.get(weekday_ko, calendar_day_type)),
        )
        events[str(date.year)].append({
            "date": date.strftime("%Y-%m-%d"),
            "month": date.month,
            "season": SEASON_MAP[season_ko],
            "dayType": day_type,
            "smp": smp,
            "actualHours": actual_hours,
        })

    for rows in events.values():
        rows.sort(key=lambda item: item["date"])
    return events


def main():
    payload = {
        "loadProfiles": read_profiles(),
        "applianceProfiles": read_appliance_profiles(),
        "events": read_events(),
    }
    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    output = (
        "// Generated by scripts/extract-reference-data.py from the uploaded audit workbooks.\n"
        "// Do not hand-edit numeric values; update the extractor and regenerate instead.\n\n"
        "export type ReferenceSeason = \"SHOULDER\" | \"SUMMER\" | \"WINTER\";\n"
        "export type ReferenceDayType = \"WEEKDAY\" | \"SATURDAY\" | \"HOLIDAY\";\n"
        "export interface ReferenceEvent {\n"
        "  date: string; month: number; season: ReferenceSeason; dayType: ReferenceDayType;\n"
        "  smp: number[]; actualHours: number[];\n"
        "}\n\n"
        f"export const REFERENCE_DATA = {body} as const;\n"
    )
    OUTPUT.write_text(output, encoding="utf-8")
    print(f"wrote {OUTPUT} ({OUTPUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
